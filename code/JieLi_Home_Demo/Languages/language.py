#!/usr/bin/python3

import openpyxl
import os
import glob
import re
from collections import OrderedDict
import pprint

KEY_PATTERN = re.compile(r'^"([^"]+)"\s*=\s*"((?:\\"|[^"])*)";\s*$')

LANGUAGE_MAP = {
    "英语(en)": "en.lproj",
    "中文(zh)": "zh-Hans.lproj",
    "日语(ja)": "ja.lproj",
    # 添加更多语言...
}

def debug_print(title, data):
    """调试信息输出"""
    print(f"\n=== DEBUG: {title} ===")
    if isinstance(data, dict):
        pprint.pprint(dict(list(data.items())[:3]))  # ✅ 正确调用方式
    else:
        print(repr(data)[:200])  # 截断长文本
    print("=" * 30)

def parse_localized_file(file_path):
    """解析本地化文件，返回（非键值行列表，键值对字典）"""
    print(f"\n[解析文件] 开始处理：{file_path}")
    
    non_key_lines = []    # 保存注释、空行等非键值对内容
    content_dict = OrderedDict()  # 保留键值对顺序
    
    if os.path.exists(file_path):
        with open(file_path, "r", encoding="utf-8") as f:
            for line_num, line in enumerate(f, 1):
                stripped_line = line.strip()
                if not stripped_line:
                    non_key_lines.append(line)
                    continue
                
                match = KEY_PATTERN.match(stripped_line)
                if match:
                    key, value = match.groups()
                    print(f"解析到键值对：{key} => {value}")
                    content_dict[key] = line  # 保留原始格式（含缩进、换行符）
                else:
                    non_key_lines.append(line)  # 注释、非法行等
    else:
        print("文件不存在，将创建新文件")
    
    debug_print("解析结果-非键值对内容", non_key_lines)
    debug_print("解析结果-键值对", content_dict)
    
    return non_key_lines, content_dict

def update_localized_file(target_path, new_content):
    """智能合并本地化文件"""
    print(f"\n[更新文件] 目标路径：{target_path}")
    
    # 解析现有内容
    non_key_lines, content_dict = parse_localized_file(target_path)
    debug_print("现有内容-非键值对内容", content_dict)
    
    # 解析新内容
    new_entries = {}
    for line_num, line in enumerate(new_content.split('\n'), 1):
        line = line.rstrip('\n')  # 去掉末尾换行符
        if not line.strip():
            continue
        match = KEY_PATTERN.match(line.strip())
        if match:
            key, value = match.groups()
            print(f"新内容键值对[{line_num}]: {key} => {value}")
            new_entries[key] = line + '\n'  # 保留原始格式
        else:
            print(f"忽略非法行[{line_num}]: {line}")
    
    # 合并内容：新内容覆盖原内容
    for key, line in new_entries.items():
        content_dict[key] = line  # 存在则替换，不存在则添加
    
    # 构建最终内容
    final_lines = non_key_lines  # 保留原注释、非键值对内容
    final_lines.extend(content_dict.values())  # 添加/替换后的键值对
    
    # 写入文件
    try:
        print(f"尝试写入文件：{target_path}")
        with open(target_path, "w", encoding="utf-8") as f:
            f.writelines(final_lines)
            
        print("文件写入成功")
    except Exception as e:
        print(f"写入文件失败：{str(e)}")
        raise

def get_brand_model_dict():
    """从Excel生成最新语言文件"""
    print("\n[生成语言文件] 开始处理Excel")
    languages_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(languages_dir, "language.xlsx")
    print(f"Excel路径：{excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["语言包"]
        print(f"工作表最大列：{ws.max_column}, 最大行：{ws.max_row}")
        
        # 按列处理语言
        for col in range(2, ws.max_column + 1):
            lang = ws.cell(row=1, column=col).value
            output_path = os.path.join(languages_dir, f"{lang}.txt")
            print(f"\n处理语言：{lang} => {output_path}")
            
            unique_entries = OrderedDict()
            for row in range(2, ws.max_row + 1):
                brand = ws.cell(row=row, column=3).value
                model = ws.cell(row=row, column=col).value
                
                print(f"行{row}：品牌[{brand}] 型号[{model}]")
                
                if not brand or not model:
                    print(f"行{row} 数据不完整，跳过")
                    continue
                
                # 动态处理重复键
                base_key = brand
                counter = 1
                while True:
                    current_key = f'"{base_key}"' if counter == 1 else f'"{base_key}_{counter}"'
                    if current_key not in unique_entries:
                        break
                    counter += 1
                print(f"生成唯一键：{current_key}")
                
                unique_entries[current_key] = f'{current_key} = "{model}";\n'
            
            # 写入临时文件
            print(f"写入临时文件：{output_path}")
            with open(output_path, "w", encoding="utf-8") as f:
                f.writelines(unique_entries.values())
            print(f"生成成功，共{len(unique_entries)}条记录")
            
    except Exception as e:
        print(f"处理Excel失败：{str(e)}")
        raise

def sync_localization():
    """同步所有语言包"""
    print("\n[同步开始]")
    languages_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(languages_dir)
    print(f"项目根目录：{project_root}")
    
    # 处理每个生成的语言文件
    for lang_file in glob.glob(os.path.join(languages_dir, "*.txt")):
        print(f"\n处理语言文件：{lang_file}")
        
        with open(lang_file, "r", encoding="utf-8") as f:
            new_content = f.read()
            print(f"读取到{len(new_content.splitlines())}行新内容")
        
        lang_name = os.path.splitext(os.path.basename(lang_file))[0]
        target_lan_name = LANGUAGE_MAP.get(lang_name)
        lproj_dir = os.path.join(project_root, target_lan_name)
        strings_path = os.path.join(lproj_dir, "Localizable.strings")
        
        print(f"目标路径：{strings_path}")
        print(f".lproj目录是否存在：{os.path.exists(lproj_dir)}")
        
        # 执行智能合并
        try:
            update_localized_file(strings_path, new_content)
        except Exception as e:
            print(f"合并失败：{str(e)}")
            continue
    
    print("\n[同步完成]")

if __name__ == "__main__":
    try:
        get_brand_model_dict()
        sync_localization()
    except Exception as e:
        print(f"\n!!! 严重错误：{str(e)}")
        raise
